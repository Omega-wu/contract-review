import copy
import os
import time
import copy

import requests
import pypdfium2 as pdfium
import json
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image as PILImage
import io

from typing import Dict, Iterator, List, Optional, Union, Any
from review.agent.parallel_doc_qa_member import ParallelDocQAMember
from review.agent.agent import Agent
from review.llm.base import BaseChatModel
from review.llm.schema import DEFAULT_SYSTEM_MESSAGE, Message, ASSISTANT
from review.utils.file_classifier import FileClassifier
from review.utils.log import logger
from review.utils.settings import PARSER_SUPPORTED_FILE_TYPES
from review.utils.parallel_executor import parallel_exec
from review.utils.settings import DEFAULT_WORKSPACE
from review.utils.utils import (extract_files_from_messages, get_file_type, hash_sha256, save_url_to_local_work_dir,
                                is_http_url)
from review.llm.schema import ContentItem
from review.configs import CONFMAP
from review.utils.excel_parser import excel_to_json
from review.utils.process_point import process_audit_point
import subprocess

OCR_URL = CONFMAP.get("OCR_URL")
SPLIT_LENGTH = CONFMAP.get("split_length")
MAX_NO_RESPONSE_RETRY = 0


class ParallelDocQA(Agent):

    def __init__(self,
                 llm: Optional[Union[Dict, BaseChatModel]] = None,
                 system_message: Optional[str] = DEFAULT_SYSTEM_MESSAGE,
                 ):

        super().__init__(
            llm=llm,
            system_message=system_message
        )
        self.data_root = os.path.join(DEFAULT_WORKSPACE, 'temporary')

    def _get_files(self, messages: List[Message]):
        session_files = extract_files_from_messages(messages, include_images=False)
        valid_files = []
        for file in session_files:
            f_type = get_file_type(file)
            if f_type in PARSER_SUPPORTED_FILE_TYPES and file not in valid_files:
                valid_files.append(file)
        return valid_files

    def _parser_json(self, content, retry_count=0):
        max_retries = 1
        json_content = copy.deepcopy(content).strip()
        if json_content.startswith('```json'):
            json_content = json_content[len('```json'):]
        if json_content.endswith('```'):
            json_content = json_content[:-3]
        try:
            content_dict = json.loads(json_content)
            return True, content_dict
        except Exception as e:
            print(e)
            if retry_count >= max_retries:
                # 达到最大重试次数，返回错误
                return False, content
            retry_content = f"""你是一个json格式修复专家，根据报错信息，将错误json转换成正确的json格式，键值对的内容务必不要改变。严禁在JSON对象之外添加任何解释、推理性内容
        
            报错信息: {e}
            原始json内容：{content}
            修复之后的json：
            """
            print(retry_content)
            error_meaasages = [{"role": "user", "content": retry_content}]
            *_, last = self._call_llm(messages=error_meaasages, stream=False)
            print(last)
            # last[-1].content
            return self._parser_json(last[-1]["content"], retry_count + 1)

    def load_pdf(self, pdf_path, dpi=144):
        images_path = []
        print("pdf_path==========", pdf_path)
        pdf = pdfium.PdfDocument(pdf_path)
        dir_name = os.path.dirname(pdf_path)
        base_name = os.path.basename(pdf_path).replace(".pdf", "")
        for index, page in enumerate(pdf, 1):
            image = self.load_pdf_page(page, dpi)
            image_path = os.path.join(dir_name, f"{base_name}_{index}.png")
            image.save(image_path)
            images_path.append(image_path)

        return images_path

    def load_pdf_page(self, page, dpi):
        scale = dpi / 72.0
        bitmap = page.render(
            scale=scale,
            rotation=0
        )
        pil_image = bitmap.to_pil()

        if pil_image.width > 3000 or pil_image.height > 3000:
            bitmap = page.render(
                scale=1.0,
                rotation=0
            )
            pil_image = bitmap.to_pil()

        return pil_image

    # ORC处理pdf
    def orc_pdf(self, pdf_path):
        url = OCR_URL
        file_info = {
            "file_id": "test_file_001"
        }
        files = {
            'file': open(pdf_path, 'rb'),
            'file_info': (None, json.dumps(file_info))
        }
        print(files)
        try:
            response = requests.post(url, files=files)
            response.raise_for_status()
            result = response.json()

            # 保存结果到JSON文件
            output_path = pdf_path.replace('.pdf', '_result.json')
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2)

            print(f"OCR处理完成，结果已保存到: {output_path}")
            return result
        except Exception as e:
            print(f"OCR处理失败: {str(e)}")
            return False

    def _run(self, messages: List[Message], **kwargs) -> [List[Message]]:

        messages = copy.deepcopy(messages)

        valid_files = self._get_files(messages)

        data = []
        all_results = []
        for file_path in valid_files:
            # file_mapping[idx] = file_path
            if is_http_url(file_path):
                # download online url
                tmp_file_root = os.path.join(self.data_root, hash_sha256(file_path))
                os.makedirs(tmp_file_root, exist_ok=True)
                file_path = save_url_to_local_work_dir(file_path, tmp_file_root)

            file_type = get_file_type(file_path)
            if file_type in ["doc", "docx"]:
                self.convert_doc_to_pdf(file_path)
                file_path = file_path.replace(f".{file_type}", ".pdf")

            # 使用requests模块调用ocr服务，去看ocr的接口和返回值
            # 文件路径
            if file_type in ["doc", "docx", "pdf"]:
                ocr_result = self.orc_pdf(file_path)

                # file_path = Path('D:/shixi/wenjian_result.json')
                # ocr_result=json.loads(file_path.read_text(encoding='utf-8'))
                print("解析结果：", ocr_result)
                # 解析excel文件，使用pandas模块读excel文件，将内容按照行转换成json格式[{"column1": "xx"， “column2”: "xx"}, {...}, {..}]
                # 将解析代码写在 utils/excel_parser.py，这里进行调用
                excel_path = r"./赛道十审核点.xlsx"
                json_result = excel_to_json(excel_path)
                #print(json_result)
                # 遍历上一步返回的内容，通过文件名字段，把属于该文件的审核点的所有行取出来，格式应该为：[{}, {}, {}...]
                # 得到的文件名字确认
                for i, item in enumerate(ocr_result):
                    # 这里根据实际来定页面、坐标信息等
                    page_content = item.get("page_content", "")
                    chunk_type = item.get("metadata", {}).get("type", "").strip()
                    if chunk_type in ["paragraph_title", "doc_title", "table_title"] and "作业票" in page_content:
                        break
                target_filename = page_content
                print(f"目标文件名: {target_filename}")


                # 使用文件分类器对文件名进行分类
                classifier = FileClassifier()
                file_classification = classifier.classify_filename(target_filename)
                print(f"文件分类结果: {file_classification['category']}")

                # 过滤属于该文件的审核点,按文件分类
                filtered_data = []
                for row in json_result:
                    # 检查文档类型匹配（如果审核点中有文档类型信息）
                    doc_type_match = False  # 默认匹配
                    if '作业类型' in row and file_classification['category'] != '未分类':
                        doc_type_match = row['作业类型'] == file_classification['category']

                    if doc_type_match:
                        filtered_data.append(row)

                print(f"找到 {len(filtered_data)} 条相关审核点")
                for index, audit_point in enumerate(filtered_data):
                    processed_point = process_audit_point(audit_point, ocr_result)
                    audit_point = processed_point["audit_point"]
                    processed_content = processed_point["processed_content"]
                    extraction_type = processed_content.get("extraction_type")
                    if extraction_type == "table_fields":
                        if not processed_content.get("html_results") and not processed_content.get("page_nums"):
                            print("没有找到对应的表格")
                            print(processed_content)
                            print(json.dumps(audit_point, ensure_ascii=False))
                            with open("loss_review.txt", "a+") as f:
                                f.write(json.dumps(audit_point, ensure_ascii=False))
                                f.write("\n\n")
                            continue
                        html_result = processed_content.get("html_results")[0]
                        print("processed_content:========================", processed_content)
                        page_num = processed_content.get("page_nums")[0]
                        prompt = audit_point["prompt"]  # 表格里填写的
                        review = audit_point["review"]
                        # prompt = prompt.replace("{ocr}", html_result)
                        prompt = prompt.replace("{review}", review)
                        images_path_list = self.load_pdf(file_path)
                        prompt_image_path = images_path_list[int(page_num)]
                        messages_new = copy.deepcopy(messages)
                        messages_new[0].content.append(ContentItem(image=prompt_image_path))
                        data.append(
                            {
                                "index": index,
                                "messages": messages_new,
                                "prompt": prompt,
                                "task": "review_table",
                                "file_path": prompt_image_path,
                                "audit_point": audit_point,
                                "target_filename": target_filename
                            }
                        )

                    elif extraction_type == "table_name_with_keywords":
                        review_res = processed_content.get("审核结果")

                        evidence = processed_content.get("审核依据")
                        page_num = processed_content.get("page_num")

                        parser_json_content = {
                            "审核结果": review_res,
                            "审核依据": evidence
                        }
                        prompt_image_path = ''
                        if page_num:
                            images_path_list = self.load_pdf(file_path)
                            prompt_image_path = images_path_list[int(page_num)]

                        result_data = {
                            'pic_path': prompt_image_path,
                            'parser_json': parser_json_content,
                            'audit_point': audit_point,
                            'target_filename': target_filename
                        }
                        all_results.append(result_data)

        logger.info('Parallel Member Num: ' + str(len(data)))

        time1 = time.time()
        results = parallel_exec(self._ask_member_agent, data, jitter=0.5)
        # results = serial_exec(self._qa, data)
        time2 = time.time()
        logger.info(f'Finished type parallel_exec. Time spent: {time2 - time1} seconds.')
        ordered_results = sorted(results, key=lambda x: x[0])

        for index, task, pic_path, text, audit_point, target_filename in ordered_results:
            parser_success, parser_json_content = self._parser_json(text)
            print(f"index:{index}\ntask:{task}\nfile_path:{file_path}\ntext:{text}")
            if parser_success:
                # 收集结果用于Excel
                result_data = {
                    'pic_path': pic_path,
                    'parser_json': parser_json_content,
                    'audit_point': audit_point,
                    'target_filename': target_filename
                }
                all_results.append(result_data)
            else:
                print("*********json解析失败***********")
                print(parser_json_content)
                print("\n\n")
                with open("error_json.txt", 'a+') as f:
                    f.write(parser_json_content)
                    f.write("\n\n")

        # 保存结果到Excel
        self.write_to_excel_with_images(all_results, "审核结果.xlsx")

        content = []
        content.extend([ContentItem(text=json.dumps(all_results, ensure_ascii=False))
                        ])

        response = [Message(role=ASSISTANT, content=content)]
        yield response

    def _ask_member_agent(self,
                          index: int,
                          messages: List[Message],
                          prompt: str,
                          task: str,
                          file_path: str,
                          audit_point: dict,
                          target_filename: str
                          ) -> tuple:
        """
           并行执行单个成员代理任务的方法。

           参数:
               index (int): 任务索引，用于标识和排序结果。
               messages (List[Message]): 消息列表，包含历史或上下文信息。
               prompt (str): 提示文本，用于指导代理处理任务。
               task (str): 任务描述，说明当前处理的具体内容。
               file_path (str): 文件路径，指向需要处理的文档或图像。

           返回:
               tuple: 包含索引、任务、文件路径和代理生成文本的元组。
           """
        doc_qa = ParallelDocQAMember(llm=self.llm)
        *_, last = doc_qa.run(messages=messages, prompt=prompt, task=task)
        # 返回元组，包含索引、任务描述、文件路径和代理返回的最终文本内容
        return index, task, file_path, last[-1].content, audit_point, target_filename

    def write_to_excel_with_images(self, data_list, output_file="审核结果.xlsx", img_width=300, img_height=300):
        """
        将包含审核结果的字典列表写入Excel文件，并嵌入图片

        Args:
            data_list: 包含审核结果的字典列表
            output_file: 输出的Excel文件名
            img_width: 图片宽度（像素）
            img_height: 图片高度（像素）
        """
        from openpyxl.styles import Alignment, Border, Side, Font

        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "审核结果"

        # 设置表头 - 增加图片路径列
        # 增加错误类型
        headers = ['文件名称', '表格名称', '审查点', '作业类型', '审核结果', '错误类型', '审核依据', '图片路径', '图片信息']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            # 设置表头样式
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # 定义边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 定义字体样式
        data_font = Font(size=12)

        # 定义对齐方式（自动换行）
        wrap_alignment = Alignment(vertical='center', wrap_text=True)

        # 填充数据
        for row_idx, item in enumerate(data_list, 2):
            # 提取所需字段
            file_name = item.get('target_filename', "")
            table_name = item.get('audit_point', {}).get('table_name', '')
            review = item.get('audit_point', {}).get('review', '')
            job_type = item.get('audit_point', {}).get('作业类型', '')
            review_res = item.get('parser_json', {}).get('审核结果', '')
            error_type = item.get('audit_point', {}).get('错误类型', '')
            review_evidence = item.get('parser_json', {}).get('审核依据', '')
            pic_path = item.get('pic_path', '')

            # 写入文本数据并设置样式
            data_cells = [
                ws.cell(row=row_idx, column=1, value=file_name),
                ws.cell(row=row_idx, column=2, value=table_name),
                ws.cell(row=row_idx, column=3, value=review),
                ws.cell(row=row_idx, column=4, value=job_type),
                ws.cell(row=row_idx, column=5, value=review_res),
                ws.cell(row=row_idx, column=6, value=error_type),  # 错误类型
                ws.cell(row=row_idx, column=7, value=review_evidence),
                ws.cell(row=row_idx, column=8, value=pic_path)  # 图片路径列
            ]

            # 为所有数据单元格设置样式
            for cell in data_cells:
                cell.font = data_font
                cell.alignment = wrap_alignment
                cell.border = thin_border

            # 设置图片信息单元格（第8列）
            pic_info_cell = ws.cell(row=row_idx, column=9)
            pic_info_cell.font = data_font
            pic_info_cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)
            pic_info_cell.border = thin_border

            # 尝试插入图片
            if pic_path and os.path.exists(pic_path):
                try:
                    # 调整图片大小
                    pil_img = PILImage.open(pic_path)
                    pil_img.thumbnail((img_width, img_height), PILImage.Resampling.LANCZOS)

                    # 转换为BytesIO对象
                    img_byte_arr = io.BytesIO()
                    pil_img.save(img_byte_arr, format=pil_img.format if pil_img.format else 'PNG')
                    img_byte_arr.seek(0)

                    # 创建openpyxl Image对象
                    img = Image(img_byte_arr)

                    # 设置图片位置（H列对应图片信息）
                    img.anchor = f'I{row_idx}'

                    # 添加到工作表
                    ws.add_image(img)

                    # 调整行高以适应图片
                    ws.row_dimensions[row_idx].height = max(80, img_height * 0.75)

                except Exception as e:
                    print(f"插入图片失败 {pic_path}: {e}")
                    pic_info_cell.value = f"图片加载失败: {os.path.basename(pic_path)}"
            else:
                pic_info_cell.value = "无图片" if not pic_path else f"图片不存在: {os.path.basename(pic_path)}"

        # 调整列宽（增加宽度确保内容完全显示）
        column_widths = {
            'A': 25,  # file_name
            'B': 25,  # table_name
            'C': 40,  # review
            'D': 20,  # 作业类型
            'E': 15,  # 审核结果
            'F': 20,  # 错误类型
            'G': 50,  # 审核依据
            'H': 40,  # 图片路径
            'I': 25,  # 图片信息
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 保存Excel文件
        wb.save(output_file)
        print(f"数据已成功写入到 {output_file}")
        print(f"共处理 {len(data_list)} 条记录")

        return output_file

    def convert_doc_to_pdf(self, doc_path, output_dir=None):
        """
        使用LibreOffice命令行将Word文档转换为PDF
        :param doc_path: 输入的Word文档路径(.doc或.docx)
        :param output_dir: 输出的PDF目录，默认为Word文档所在目录
        :return: 转换成功返回True，否则返回False
        """
        doc_extension = os.path.splitext(doc_path)[1]
        print(doc_extension)
        if doc_extension == ".pdf":
            return True
        if doc_extension in [".doc", ".docx"]:
            # 检查输入文件是否存在
            if not os.path.exists(doc_path):
                print(f"错误：文件不存在 - {doc_path}")
                return False

            # 如果未指定输出目录，则使用输入文件所在目录
            if output_dir is None:
                output_dir = os.path.dirname(doc_path)
            # 确保输出目录存在
            os.makedirs(output_dir, exist_ok=True)

            try:
                # 构建命令
                # 注意：在某些系统上，命令可能是 'libreoffice7.0' 等带版本号的形式[8](@ref)
                cmd = [
                    'libreoffice',  # 或 'libreoffice7.0', 'soffice'
                    '--headless',  # 无界面模式
                    '--convert-to',
                    'pdf',
                    '--outdir',
                    output_dir,
                    doc_path
                ]

                # 执行转换命令
                subprocess.run(cmd, check=True, capture_output=True, text=True, timeout=60)
                print(f"转换成功！PDF文件保存在：{output_dir}")
                return True

            except subprocess.CalledProcessError as e:
                print(f"转换失败，LibreOffice返回错误：{e.stderr}")
                return False
            except subprocess.TimeoutExpired:
                print("转换超时")
                return False
            except FileNotFoundError:
                print("未找到LibreOffice，请确保已安装并将其添加到系统PATH环境变量中")
                return False
            except Exception as e:
                print(f"转换过程中发生未知错误：{e}")
                return False
